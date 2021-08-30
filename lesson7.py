# Author : 柠檬班-亚萌
# Project : scb23
# Time : 2021/8/26 20:02
# E-mail : 3343787213@qq.com
# Company : 湖南零檬信息技术有限公司
# Site : http://www.lemonban.com
# Forum : http://testingpai.com

'''
1、编写自动化测试用例，代码自动读取数据   # read_data()
2、发送请求，得到响应结果   # func()
3、执行结果（响应结果）  VS  预期结果
4、写入最终的真实结果到测试用例  # write_data()
'''
import requests
import jsonpath
import openpyxl


def read_data(filename, sheetname):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    max_row = sheet.max_row  # 获取总行数
    cases_list = [] # 定义一个空列表，用来存储测试数据
    for i in range(2, max_row+1, 1): # 取头不取尾\
        dict1 =dict(
        id = sheet.cell(row=i, column=1).value,
        header = sheet.cell(row=i, column=5).value,  # 获取请求头
        url = sheet.cell(row=i, column=6).value,  # 获取接口地址
        body = sheet.cell(row=i, column=7).value, # 获取请求体
        expect = sheet.cell(row=i, column=8).value) # 获取预期结果
        cases_list.append(dict1)  # 把每一条测试数据添加到列表里保存
    return cases_list

def func(url,body,header):
    res = requests.post(url=url, json=body, headers=header)
    res_res = res.json()
    return res_res

def write_data(filenanme,sheetname,row, column, final_result):
    wb = openpyxl.load_workbook(filenanme)
    sheet = wb[sheetname]
    sheet.cell(row=row, column=column).value = final_result  # 写入执行结果
    wb.save(filenanme)


def execute_func(filename,sheetname):
    cases = read_data(filename,sheetname)  # 调用读取函数，将测试数据取出来
    for case in cases:
        # excel里面读取出来的数据，都是str类型
        id = case.get('id') # 取出id
        # id = case['id']
        url = case.get('url') # 取出接口地址
        header = case.get('header') # 取出请求头
        body = case.get('body') # 取出请求体
        expect = case.get('expect') # 取出预期结果
        header = eval(header)  #运行被字符串包裹着的python表达式，然后把引号去掉，取字符串里面的表达式
        body = eval(body)
        expect = eval(expect)
        expect_code = expect.get('code') # 取出预期结果里面的code进行断言
        res1 = func(url=url,body=body,header=header)  # 调用发送请求的函数执行请求
        real_code = res1.get('code')
        print("预期结果的code为{}".format(expect_code))
        print("实际结果的code为{}".format(real_code))
        if expect_code == real_code:  # 做结果的判断
            print("{}功能，第{}条用例通过！！".format(sheetname,id))
            print('*' * 50)
            final_res = '通过'
        else:
            print("{}功能，第{}条用例不通过！！".format(sheetname,id))
            print('*' * 50)
            final_res = '不通过'
        write_data(filename,sheetname,id+1, 9,final_res)  # 调动写入函数，将最终结果写回到excel

execute_func('testcase_api_wuye.xlsx', 'register')
execute_func('testcase_api_wuye.xlsx', 'login')